# /bin/bash python

import openpyxl
result = openpyxl.Workbook()
result_sheet = result['Sheet']
result_sheet.title = '公司周报'
wb_list = ['开发部周报.xlsx','教学部周报.xlsx','运营部周报.xlsx']


row_num = 1
for wb_name in wb_list:
    wb = openpyxl.load_workbook(wb_name)

    for sheet in wb:
        for index,row in enumerate(sheet.rows):
            if index > 0:
                for cell in row:
                    result_sheet.cell(row=row_num,column=cell.column).value = cell.value
                row_num += 1

header = ['部门','工作内容','是否完成','延期','责任人','今日进度']
result_sheet.insert_rows(1)

for i in range(len(header)):
    result_sheet.cell(row=1,column=i+1).value = header[i]


result.save('./tmp/合并周报.xlsx')